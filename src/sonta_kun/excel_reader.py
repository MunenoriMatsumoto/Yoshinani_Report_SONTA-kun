"""Excelファイル読み込みモジュール"""

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class TodoItem:
    """ToDoアイテム"""

    task: str
    status: str = ""
    due_date: Optional[date] = None
    priority: str = ""
    category: str = ""
    notes: str = ""

    def to_dict(self) -> dict:
        """辞書形式に変換"""
        return {
            "task": self.task,
            "status": self.status,
            "due_date": self.due_date.isoformat() if self.due_date else None,
            "priority": self.priority,
            "category": self.category,
            "notes": self.notes,
        }


@dataclass
class TodoList:
    """ToDoリスト"""

    items: list[TodoItem] = field(default_factory=list)
    source_file: Optional[str] = None

    def to_dict(self) -> dict:
        """辞書形式に変換"""
        return {
            "source_file": self.source_file,
            "items": [item.to_dict() for item in self.items],
        }

    def to_text(self) -> str:
        """テキスト形式に変換（AI入力用）"""
        lines = []
        for i, item in enumerate(self.items, 1):
            line = f"{i}. {item.task}"
            if item.status:
                line += f" [{item.status}]"
            if item.due_date:
                line += f" (期限: {item.due_date})"
            if item.priority:
                line += f" 優先度: {item.priority}"
            if item.category:
                line += f" カテゴリ: {item.category}"
            if item.notes:
                line += f"\n   備考: {item.notes}"
            lines.append(line)
        return "\n".join(lines)


class ExcelReader:
    """Excelファイルからタスクリストを読み込むクラス"""

    # よく使われるカラム名のマッピング
    COLUMN_MAPPINGS = {
        "task": ["タスク", "task", "タスク名", "作業内容", "内容", "件名", "タイトル", "title"],
        "status": ["ステータス", "status", "状態", "進捗", "完了"],
        "due_date": ["期限", "due_date", "期日", "締切", "deadline", "完了予定日"],
        "priority": ["優先度", "priority", "重要度"],
        "category": ["カテゴリ", "category", "分類", "種別"],
        "notes": ["備考", "notes", "メモ", "コメント", "詳細"],
    }

    def __init__(self):
        self._column_indices: dict[str, int] = {}

    def read(self, file_path: str | Path) -> TodoList:
        """
        Excelファイルを読み込んでTodoListを返す

        Args:
            file_path: Excelファイルのパス

        Returns:
            読み込んだTodoList

        Raises:
            ExcelReadError: ファイル読み込みに失敗した場合
        """
        path = Path(file_path)

        if not path.exists():
            raise ExcelReadError(f"ファイルが見つかりません: {path}")

        if path.suffix.lower() not in [".xlsx", ".xls", ".xlsm"]:
            raise ExcelReadError(f"サポートされていないファイル形式です: {path.suffix}")

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active

            if sheet is None:
                raise ExcelReadError("アクティブなシートが見つかりません")

            items = self._parse_sheet(sheet)

            return TodoList(items=items, source_file=str(path))

        except ExcelReadError:
            raise
        except Exception as e:
            raise ExcelReadError(f"Excelファイルの読み込みに失敗しました: {e}") from e

    def _parse_sheet(self, sheet: Worksheet) -> list[TodoItem]:
        """シートをパースしてTodoItemのリストを返す"""
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return []

        # ヘッダー行を検出
        header_row_idx = self._find_header_row(rows)
        if header_row_idx is None:
            # ヘッダーがない場合は1列目をタスクとして扱う
            return self._parse_without_header(rows)

        headers = rows[header_row_idx]
        self._map_columns(headers)

        # データ行をパース
        items = []
        for row in rows[header_row_idx + 1 :]:
            item = self._parse_row(row)
            if item:
                items.append(item)

        return items

    def _find_header_row(self, rows: list[tuple]) -> Optional[int]:
        """ヘッダー行のインデックスを見つける"""
        for idx, row in enumerate(rows[:10]):  # 最初の10行を検索
            if row and any(self._is_header_cell(cell) for cell in row if cell):
                return idx
        return None

    def _is_header_cell(self, value) -> bool:
        """セルがヘッダーかどうかを判定"""
        if not isinstance(value, str):
            return False
        value_lower = value.lower().strip()
        for field_names in self.COLUMN_MAPPINGS.values():
            if value_lower in [name.lower() for name in field_names]:
                return True
        return False

    def _map_columns(self, headers: tuple) -> None:
        """ヘッダーからカラムインデックスをマッピング"""
        self._column_indices = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            header_lower = str(header).lower().strip()
            for field_name, aliases in self.COLUMN_MAPPINGS.items():
                if header_lower in [alias.lower() for alias in aliases]:
                    self._column_indices[field_name] = idx
                    break

    def _parse_row(self, row: tuple) -> Optional[TodoItem]:
        """行をパースしてTodoItemを返す"""
        task_idx = self._column_indices.get("task")

        # タスク列がない場合は最初の非空セルをタスクとする
        if task_idx is None:
            for idx, cell in enumerate(row):
                if cell and str(cell).strip():
                    task_idx = idx
                    break

        if task_idx is None or task_idx >= len(row):
            return None

        task = row[task_idx]
        if not task or not str(task).strip():
            return None

        return TodoItem(
            task=str(task).strip(),
            status=self._get_cell_value(row, "status"),
            due_date=self._get_date_value(row, "due_date"),
            priority=self._get_cell_value(row, "priority"),
            category=self._get_cell_value(row, "category"),
            notes=self._get_cell_value(row, "notes"),
        )

    def _parse_without_header(self, rows: list[tuple]) -> list[TodoItem]:
        """ヘッダーなしでパースする（1列目をタスクとして扱う）"""
        items = []
        for row in rows:
            if row and row[0] and str(row[0]).strip():
                items.append(TodoItem(task=str(row[0]).strip()))
        return items

    def _get_cell_value(self, row: tuple, field_name: str) -> str:
        """指定フィールドのセル値を取得"""
        idx = self._column_indices.get(field_name)
        if idx is None or idx >= len(row):
            return ""
        value = row[idx]
        return str(value).strip() if value else ""

    def _get_date_value(self, row: tuple, field_name: str) -> Optional[date]:
        """指定フィールドの日付値を取得"""
        idx = self._column_indices.get(field_name)
        if idx is None or idx >= len(row):
            return None
        value = row[idx]
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return None


class ExcelReadError(Exception):
    """Excel読み込みエラー"""

    pass
