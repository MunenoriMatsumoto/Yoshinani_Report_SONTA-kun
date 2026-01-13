"""Excel読み込みモジュールのテスト"""

import tempfile
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from sonta_kun.excel_reader import ExcelReadError, ExcelReader, TodoItem, TodoList


@pytest.fixture
def sample_excel_file():
    """サンプルExcelファイルを作成"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb = Workbook()
        ws = wb.active
        ws.title = "タスク"

        # ヘッダー
        ws["A1"] = "タスク"
        ws["B1"] = "ステータス"
        ws["C1"] = "期限"
        ws["D1"] = "優先度"

        # データ
        ws["A2"] = "機能Aの実装"
        ws["B2"] = "進行中"
        ws["C2"] = date(2024, 12, 31)
        ws["D2"] = "高"

        ws["A3"] = "テスト作成"
        ws["B3"] = "未着手"
        ws["C3"] = date(2025, 1, 15)
        ws["D3"] = "中"

        wb.save(f.name)
        yield Path(f.name)

    Path(f.name).unlink(missing_ok=True)


@pytest.fixture
def simple_excel_file():
    """シンプルなExcelファイルを作成（ヘッダーなし）"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb = Workbook()
        ws = wb.active

        ws["A1"] = "タスク1"
        ws["A2"] = "タスク2"
        ws["A3"] = "タスク3"

        wb.save(f.name)
        yield Path(f.name)

    Path(f.name).unlink(missing_ok=True)


def test_todo_item_to_dict():
    """TodoItemの辞書変換テスト"""
    item = TodoItem(
        task="テストタスク",
        status="完了",
        due_date=date(2024, 12, 31),
        priority="高",
    )
    result = item.to_dict()

    assert result["task"] == "テストタスク"
    assert result["status"] == "完了"
    assert result["due_date"] == "2024-12-31"
    assert result["priority"] == "高"


def test_todo_list_to_text():
    """TodoListのテキスト変換テスト"""
    items = [
        TodoItem(task="タスク1", status="完了"),
        TodoItem(task="タスク2", status="進行中", priority="高"),
    ]
    todo_list = TodoList(items=items)
    text = todo_list.to_text()

    assert "タスク1" in text
    assert "[完了]" in text
    assert "タスク2" in text
    assert "優先度: 高" in text


def test_excel_reader_with_headers(sample_excel_file):
    """ヘッダー付きExcelファイルの読み込みテスト"""
    reader = ExcelReader()
    result = reader.read(sample_excel_file)

    assert len(result.items) == 2
    assert result.items[0].task == "機能Aの実装"
    assert result.items[0].status == "進行中"
    assert result.items[0].due_date == date(2024, 12, 31)
    assert result.items[0].priority == "高"


def test_excel_reader_without_headers(simple_excel_file):
    """ヘッダーなしExcelファイルの読み込みテスト"""
    reader = ExcelReader()
    result = reader.read(simple_excel_file)

    assert len(result.items) == 3
    assert result.items[0].task == "タスク1"
    assert result.items[1].task == "タスク2"


def test_excel_reader_file_not_found():
    """存在しないファイルのエラーテスト"""
    reader = ExcelReader()

    with pytest.raises(ExcelReadError, match="ファイルが見つかりません"):
        reader.read("/nonexistent/file.xlsx")


def test_excel_reader_unsupported_format():
    """サポートされていないファイル形式のエラーテスト"""
    with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as f:
        f.write(b"test")
        temp_path = Path(f.name)

    try:
        reader = ExcelReader()
        with pytest.raises(ExcelReadError, match="サポートされていないファイル形式"):
            reader.read(temp_path)
    finally:
        temp_path.unlink(missing_ok=True)
